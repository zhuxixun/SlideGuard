"""Excel report generation (.xlsx with issue list)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from backend.models.issue import ScanResult


def generate_excel_report(result: ScanResult, output_path: str):
    """Generate Excel report with issue summary and detail."""
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws_summary = wb.active
    ws_summary.title = "概览"

    header_font = Font(bold=True, size=14, color="1F4E79")
    ws_summary.cell(row=1, column=1, value="SlideGuard 质检报告").font = header_font

    info_data = [
        ("文件名", result.file_name),
        ("页数", result.total_pages),
        ("总体评分", result.score),
        ("问题总数", result.total_issues),
        ("S1严重问题", result.s1_count),
        ("S2高风险问题", result.s2_count),
        ("S3一般问题", result.s3_count),
        ("S4建议优化", result.s4_count),
        ("可自动修复", result.fixable_count),
        ("页面尺寸", f"{result.page_size[0]:.1f}x{result.page_size[1]:.1f} 英寸"),
    ]

    for i, (label, value) in enumerate(info_data):
        row = i + 3
        ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=value)

    # Dimension scores
    if result.dimension_scores:
        row = len(info_data) + 5
        ws_summary.cell(row=row, column=1, value="各维度得分").font = Font(bold=True, size=12)
        for dim, score in result.dimension_scores.items():
            row += 1
            ws_summary.cell(row=row, column=1, value=dim)
            ws_summary.cell(row=row, column=2, value=score)

    # --- Sheet 2: Issues ---
    ws_issues = wb.create_sheet("问题清单")

    headers = ["编号", "页面", "严重程度", "规则", "问题描述", "实际值", "期望值", "建议", "可自动修复"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")

    for col, header in enumerate(headers, 1):
        cell = ws_issues.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill

    severity_fills = {
        "S1": PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"),
        "S2": PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"),
        "S3": PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid"),
        "S4": PatternFill(start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"),
    }

    for i, issue in enumerate(result.all_issues):
        row = i + 2
        ws_issues.cell(row=row, column=1, value=i + 1)
        ws_issues.cell(row=row, column=2, value=f"第{issue.location.slide_index + 1}页")
        ws_issues.cell(row=row, column=3, value=issue.severity)
        ws_issues.cell(row=row, column=4, value=issue.rule_name)
        ws_issues.cell(row=row, column=5, value=issue.description)
        ws_issues.cell(row=row, column=6, value=issue.actual_value or "")
        ws_issues.cell(row=row, column=7, value=issue.expected_value or "")
        ws_issues.cell(row=row, column=8, value=issue.suggestion)
        ws_issues.cell(row=row, column=9, value="是" if issue.auto_fixable else "否")

        # Apply severity coloring
        fill = severity_fills.get(issue.severity)
        if fill:
            for col in range(1, 10):
                ws_issues.cell(row=row, column=col).fill = fill

    # Auto-adjust column widths
    for ws in [ws_summary, ws_issues]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 60)

    wb.save(output_path)
    return output_path
